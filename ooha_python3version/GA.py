# coding: utf-8

# In[ ]:


import random
import math
import numpy
import dill
from itertools import product
from collections import OrderedDict
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os

fold = os.getcwd()
path = fold + "/realrundata_ga.xlsx"
column_Name = ('Min_Ga','Sum_Ga','Time')
if os.path.exists("realrundata_ga.xlsx"):
    wb = load_workbook(filename= path)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active
    sheet.append(column_Name)
    
re = []

def Genetic(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum, Conflict,tStart):
    alive = 45      #每代保留基因數
    child = 50     #種群數量
    t = 600        #迭代次數
    mutent = 0.02   #突變機率

    #種群根據適應函數排列大小
    def sort_by_obj(mother,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum):
        score = []
        for i in mother:
            score.append(obj_value(i,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum))
        for j in range(len(score)):
            for k in range(j+1,len(score)):
                if(score[j]<score[k]):
                    t = score[j]
                    score[j] = score[k]
                    score[k] = t
                    g = mother[j]
                    mother[j] = mother[k]
                    mother[k] = g
        return mother

    def validCheck(y,TotalN,K,M,AdTime, Conflict):
        AdRepeatTime = 10
        AdRepeatNum = 10
        AdContinousNum = 3
        ConflictNum = 5

        previous = -1 #上個廣告
        continuous = 0 #重複次數
        indi_i = 0;
        for i in range(TotalN):
            if i == indi_i:
                #廣告時間段切錯
                for j in range(1,AdTime[int(y[i])]):
                    if i+j>=TotalN:
                        break
                    if y[i+j] != y[i]:
                        #print('切錯')
                        return False
                #重複播放限制
                if y[i]==previous:
                    continuous += 1
                else:
                    continuous = 1
                if continuous > AdContinousNum:
                    #print('重複')
                    return False
                previous = y[i]
                #一段時間重複播放限制
                counter = 0
                for j in range(1,AdTime[int(y[i])]*AdRepeatTime):
                    if i+j>=TotalN:
                        break
                    if y[i+j] == y[i]:
                        counter +=1
                if counter > AdTime[int(y[i])]*AdRepeatNum:
                    #print('時間段重複')
                    return False
                #衝突限制
                for j in range(1,ConflictNum):
                    if i+j>=TotalN:
                        break
                    if Conflict[int(y[i]),int(y[i+j])] == 0:
                        #print('衝突')
                        return False
                indi_i += AdTime[int(y[i])]
        return True
        
        
    #適應函數
    def obj_value(y,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum):
        yRan = {}
        indi_T = 0
        for n in range(TotalN):
            for m in range(M):
                yRan[n,m] = 0
            if n == indi_T:
                yRan[n,y[n]] = 1
                indi_T = indi_T + AdTime[int(y[n])]
        Ad_RanPre = {}
        for j in range(M):
            Ad_RanPre[j] = sum( (sum(AudienceNum[w,k]*AudiencePreference[j,k] for k in range(K))*sum( yRan[i,j] for i in range(TotalBusTime[w], TotalBusTime[w+1]) ))for w in range(StopNum) )
        Cal2 = list(Ad_RanPre.values())
        return Ad_RanPre[min(Ad_RanPre, key = Ad_RanPre.get)]
    
    #初始種群
    mother = []
    for i in range(child):
        valid = False
        while valid == False:
            indi_T = 0
            y = numpy.zeros(TotalN)
            indi_AD = random.randint(0,M-1)
            for n in range(TotalN):
                if n == indi_T:
                    for j in range(indi_T,indi_T+AdTime[indi_AD]):
                        if(j>=TotalN):
                            break;
                        else:
                            y[j] = indi_AD
                    indi_T = indi_T + AdTime[indi_AD]
                    indi_AD = random.randint(0,M-1)
            valid = validCheck(y,TotalN,K,M,AdTime, Conflict)
        mother.append(y)
    
    #開始迭代
    tEnd = time.time()
    i = 0
    while tEnd-tStart<350:
        tmp = sort_by_obj(mother,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum)
        mother = []
        global re
        re.append(obj_value(tmp[0],K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum))
        if (i%100) == 0:
            print('the',i,'th Times ,MAX=',obj_value(tmp[0],K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum))
        for j in range(alive):
            mother.append(tmp[j])
        for j in range(child-alive):      
            valid = False
            while valid == False:
                #選擇要交配的基因
                parent1 = random.randint(0,alive-1)
                parent2 = random.randint(0,alive-1)
                while parent1 ==parent2:
                    parent1 = random.randint(0,alive-1)
                    parent2 = random.randint(0,alive-1)
                #選擇交配位置 
                section = random.randint(0,int(TotalN/3))*3
                new = []
                for k in range(TotalN):
                    if(k<section):
                        new.append(mother[parent1][k])
                    else:
                        new.append(mother[parent2][k])
                #決定是否突變
                if(random.random()<mutent):
                    mut = random.randint(0,int(TotalN/3))*3
                    mutAd = random.randint(0,M-1)
                    for w in range(3):
                        if(mut+w>=TotalN):
                            break
                        new[mut+w] = mutAd
                valid = validCheck(new,TotalN,K,M,AdTime, Conflict)
            mother.append(new)
        tEnd = time.time()
        i = i + 1

    yRan = {}
    indi_T = 0
    for n in range(TotalN):
        for m in range(M):
            yRan[n,m] = 0
        if n == indi_T:
            yRan[n,tmp[0][n]] = 1
            indi_T = indi_T + AdTime[int(tmp[0][n])]
    Ad_RanPre = {}
    for j in range(M):
        Ad_RanPre[j] = sum( (sum(AudienceNum[w,k]*AudiencePreference[j,k] for k in range(K))*sum( yRan[i,j] for i in range(TotalBusTime[w], TotalBusTime[w+1]) ))for w in range(StopNum) )
    Cal2 = list(Ad_RanPre.values())
    return tmp[0],Ad_RanPre[min(Ad_RanPre, key = Ad_RanPre.get)], numpy.sum(Cal2)

def DataLoader(caseNum):
    file = open('data','rb')
    data = dill.load(file)
    timeResult = data[caseNum][12]
    #StopNum：站數
    StopNum = len(timeResult)
    #Nw：第w站所花時間段數量，maxN：各站時間段總數
    N = {}
    temp = 0
    TotalN = 0
    TickSize=10
    TotalBusTime = {}
    TotalBusTime[0] = 0
    for w in range(StopNum-1):
        N[w] = math.ceil((timeResult[w+1]-timeResult[w]-temp)/TickSize)
        TotalN = TotalN + N[w]
        TotalBusTime[w+1] = TotalN
        temp = TotalBusTime[w+1]*TickSize-timeResult[w+1]
    N[StopNum-1] = 0
    TotalN = TotalN + N[StopNum-1]
    TotalBusTime[StopNum] = TotalN

    at = load_workbook('adtime.xlsx')
    #AdTime j: 第j種廣告時長會佔據多少時間段
    AdTime = {}
    AdName = {}
    for j in range(at['1'].max_row-1):
        AdTime[j] = at['1']['B'][j+1].value
        AdName[at['1']['A'][j+1].value] = j
    #M：廣告數量
    M = at['1'].max_row-1
    #AdRepeatTime, AdRepeatNum: 最短P個時間段內不可播放同樣廣告超過K次
    AdRepeatTime = 10
    AdRepeatNum = 10
    #AdContinousNum: 不可連續撥放AdContinousNum次
    AdContinousNum = 5

    #Conflict j1j2, ConflictNum: 第j1和第j2個廣告可同路線ConflictNum次內播放(1:可    0:不可)
    Conflict = {}
    for j1 in range(M):
        for j2 in range(M):
            Conflict[j1,j2] = 1
    cft = load_workbook('conflict.xlsx')
    for i in range(cft['1'].max_row):
        j1 = AdName.get(cft["1"]["A"][i].value)
        j2 = AdName.get(cft["1"]["B"][i].value)
        Conflict[j1,j2] = 0
        Conflict[j2,j1] = 0

    ConflictNum = 5
    #AudiencePreference jk: 客群k喜歡廣告j的程度 
    AudiencePreference = {}
    pre = load_workbook('preference.xlsx')
    for j in range(M):
        temp = 0
        for k in pre['1'][j+2]:
            if temp == 0:
                j1 = AdName.get(k.value)
                temp = temp+1
            else:
                AudiencePreference[j1,temp-1] = k.value 
                temp = temp+1
    #K: 客群種類數
    K = temp-1
    passengerResults = {}
    for i in range(12):
        passengerResults[i] = data[caseNum][i]
    #AudienceNum w,k: w時段車上有幾位屬於客群k的人
    AudienceNum = {}
    for k in range(K):
        for w in range(StopNum):
            AudienceNum[w,k] = data[caseNum][k][w]
        
    for i in range(M):
        AdTime[i] = int(AdTime[i]/TickSize)
    tStart = time.time()
    result,G_obj,G_sum = Genetic(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum,Conflict,tStart)
    tEnd = time.time()
    ans = (G_obj,G_sum,tEnd-tStart)
    print('排程結果:',result,'最小曝光度:',G_obj,'總合曝光度:',G_sum)
    return ans

for i in range(0,25): #改這邊的range，上限99
    data = DataLoader(i)
    sheet.append(data)
    wb.save(path)