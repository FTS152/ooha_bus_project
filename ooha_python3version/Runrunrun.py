
# coding: utf-8

# In[7]:


import random
import math
import numpy
from itertools import product
from collections import OrderedDict
from gurobipy import *
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os

Scene_AD = list(range(3))
Scene_Slot = list(range(3))
Scene_Pass = list(range(3))
Scene_Pre = list(range(4))
Scene_Conflict = list(range(3))
custom = [list(i) for i in product(Scene_Slot,Scene_AD, Scene_Pass, Scene_Pre, Scene_Conflict)]

fold = os.getcwd()
path = fold + "/data.xlsx"
column_Name = ('# of Ads','TimeSlots', 'Passenger', 'Preference', 'Conflict', 'Min_Turn','Sum_Turn','Mean_Turn','Std_Turn', 'Min_Random','Sum_Random','Mean_Random','Std_Random','Min_Obj','Sum_Obj','Mean_Obj','Std_Obj', 'Gap','Time','seed')
if os.path.exists("data.xlsx"):
    wb = load_workbook(filename= path)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active
    sheet.append(column_Name)
    
def TakeTurn(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum):
    yTurn = {}
    #計算哪個廣告開始在哪個時間段開始播
    indi_AD = 0
    indi_T = 0
    for n in range(TotalN):
        for m in range(M):
            yTurn[n,m] = 0
        if n is indi_T:
            yTurn[n,indi_AD] = 1
            indi_T = indi_T + AdTime[indi_AD]
            indi_AD = (indi_AD+1)%M
    #開始算objective
    Ad_TurnPre = {}
    for j in range(M):
        Ad_TurnPre[j] = sum( (sum(AudienceNum[w,k]*AudiencePreference[j,k] for k in range(K))*sum( yTurn[i,j] for i in range(TotalBusTime[w], TotalBusTime[w+1]) ))for w in range(StopNum) )
    Cal1 = list(Ad_TurnPre.values())
    return Ad_TurnPre[min(Ad_TurnPre, key = Ad_TurnPre.get)], numpy.sum(Cal1), numpy.mean(Cal1),numpy.std(Cal1)

def RandomTurn(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum, seed):
    random.seed(seed)
    yRan = {}
    indi_AD = 0
    indi_T = 0
    for n in range(TotalN):
        for m in range(M):
            yRan[n,m] = 0
        if n is indi_T:
            yRan[n,indi_AD] = 1
            indi_T = indi_T + AdTime[indi_AD]
            indi_AD = random.randint(0,M-1)
    Ad_RanPre = {}
    for j in range(M):
        Ad_RanPre[j] = sum( (sum(AudienceNum[w,k]*AudiencePreference[j,k] for k in range(K))*sum( yRan[i,j] for i in range(TotalBusTime[w], TotalBusTime[w+1]) ))for w in range(StopNum) )
    Cal2 = list(Ad_RanPre.values())
    return Ad_RanPre[min(Ad_RanPre, key = Ad_RanPre.get)], numpy.sum(Cal2), numpy.mean(Cal2),numpy.std(Cal2)
    
def mycallback(model, where):
    if where == GRB.Callback.MIP:
        time = model.cbGet(GRB.Callback.RUNTIME)
        objbst = model.cbGet(GRB.Callback.MIP_OBJBST)
        objbnd = model.cbGet(GRB.Callback.MIP_OBJBND)
        if time > 2400.0:
            model.terminate()
        elif time > 1800.0:
            if abs(objbst-objbnd) < 0.04 * (1.0 + abs(objbst)):
                print("Gap set to 4%")
                model.terminate()                
        elif time > 1500.0:
            if abs(objbst-objbnd) < 0.03 * (1.0 + abs(objbst)):
                print("Gap set to 3%")
                model.terminate()
        elif time > 1200.0:
            if abs(objbst-objbnd) < 0.02 * (1.0 + abs(objbst)):
                print("Gap set to 2%")
                model.terminate()
        elif time > 1500.0:
            if abs(objbst-objbnd) < 0.015 * (1.0+ abs(objbst)):
                print("Gap set to 1%")
                model.terminate()
        elif time > 600.0:
            if abs(objbst-objbnd) < 0.01 * (1.0+ abs(objbst)):
                print("Gap set to 1%")
                model.terminate()
                
def OurAlgorithm(StopNum, K,N,M, TotalN, AdTime,AudienceNum, AudiencePreference,TotalBusTime, Conflict):
    m = Model("project")
    #Parameters
    Container = 10000
    TickSize = 10
    Safety = 0
    AdRepeatTime = 10
    AdRepeatNum = 10
    AdContinousNum = 5
    ConflictNum = 5
    #Create variables
    #ans:  最大化廣告商的最低hit數
    ans=m.addVar(lb=0,vtype=GRB.CONTINUOUS,name='ans')
    x={} #xij: 第i個時間段中是否有在播放第j種廣告( 1:播放    0:沒有播放)
    for i in range(TotalN):
        for j in range(M):
            x[i,j] = m.addVar(vtype=GRB.BINARY, name='x_%d_%d'%(i,j))
    y={} #yij:第i個時間段時第j種廣告開始播放( 1:播放    0:沒有播放)
    for i in range(TotalN):
        for j in range(M):
            y[i,j] = m.addVar(vtype=GRB.BINARY, name='y_%d_%d'%(i,j))
    z={} #zj: 若第j個廣告是否會在這次規劃中播出(1:會  0:不會)
    for j in range(M):
        z[j] = m.addVar(vtype=GRB.BINARY, name='z_%d'%j)
    c={} #cij1j2: 處理衝突限制式
    for j1 in range(M):
        j2 = j1+1
        while j2<M:
            for i in range(TotalN-ConflictNum):
                c[i,j1,j2] = m.addVar(vtype=GRB.BINARY, name='c_%d_%d_%d'%(i,j1,j2))
            j2 = j2+1
    t={}
    for j in range(M):
        t[j] = m.addVar(name='t_%d'%j)
    # Integrate new variables    
    m.update()
    #Constraints
    m.setObjective( ans , GRB.MAXIMIZE ) # Set objective 最大化廣告每日的最低hit數
    # 定義yij範圍
    for j in range(M):
        for i in range(TotalN):
            m.addConstr( y[i,j] <= x[i,j] )
            m.addConstr( quicksum( y[max(i-i0,0),j] for i0 in range(AdTime[j]) ) >= x[i,j] )
            for i1 in range(AdTime[j]):
                if i+i1<TotalN: 
                    m.addConstr( y[i,j] <= x[i+i1,j] )
                    if i1 >= 1:
                        m.addConstr( y[i,j] + y[i+i1,j] <=1 )
    # 定義zj範圍
    for j in range(M):
        m.addConstr( quicksum(y[i,j] for i in range(TotalN) ) >= z[j] )
        m.addConstr( quicksum(y[i,j] for i in range(TotalN) ) <= 999999999*z[j] )

    #每個時間段只會播一種廣告
    for i in range(TotalN):
         m.addConstr( quicksum(x[i,j] for j in range(M)) <= 1 )

    #廣告總長度不超過上限 
    m.addConstr( quicksum(quicksum(x[i,j] for j in range(M)) for i in range(TotalN) ) * TickSize <= TickSize*TotalN-Safety )

    #在AdRepeatTime個時間段內最多播AdrepeatNum次
    for j in range(M):
        for i1 in range(TotalN-AdRepeatTime):
            m.addConstr( quicksum( y[i2,j] for i2 in range(i1,i1+AdRepeatTime) ) <= AdRepeatNum )

    #第j1和第j2個廣告可同路線ConflictNum次內播放
    for j1 in range(M):
        j2 = j1+1
        while j2<M:
            for i1 in range(TotalN-ConflictNum):
                m.addConstr( quicksum( y[i1,j1] for i1 in range(i1,ConflictNum) ) <= 10*TotalN*(Conflict[j1,j2] + 1-c[i1,j1,j2]) )
                m.addConstr( quicksum( y[i1,j2] for i1 in range(i1,ConflictNum) ) <= 10*TotalN*(Conflict[j1,j2] + 1-c[i1,j1,j2]) )
            j2 = j2+1    

    #AdContinousNum: 不可連續撥放AdContinousNum次
    for j in range(M):
        for i1 in range( TotalN-AdTime[j]*(AdContinousNum-1) ):
            m.addConstr( quicksum( y[i1+i2*AdTime[j],j] for i2 in range(AdContinousNum) ) <= AdContinousNum-1 )
    hit = {}
    #ans: 所有廣告中最低的廣告hit數
    for j in range(M):
        m.addConstr( quicksum( ((quicksum( AudienceNum[w,k]*AudiencePreference[j,k] for k in range(K)))* (quicksum( y[i,j] for i in range(TotalBusTime[w],TotalBusTime[w+1]) ) ) )for w in range(StopNum) ) >= ans )
        hit[j] = quicksum( ((quicksum( AudienceNum[w,k]*AudiencePreference[j,k] for k in range(K)))* (quicksum( y[i,j] for i in range(TotalBusTime[w],TotalBusTime[w+1]) ) ) )for w in range(StopNum) )
    for j in range(M):
        m.addConstr( quicksum( ((quicksum( AudienceNum[w,k]*AudiencePreference[j,k] for k in range(K)))* (quicksum( y[i,j] for i in range(TotalBusTime[w],TotalBusTime[w+1]) ) ) )for w in range(StopNum) ) ==t[j] )    

    m.setParam("MIPGap", 0.005)
    
    tstart = time.time()
    m.optimize(mycallback)
    tend = time.time()
    tpass = tend-tstart
    m.write("project.lp")
    Cal3 = []
    for i in range(M):
        Cal3.append(t[i].X)
    return m.objVal, m.MIPGap, tpass, numpy.sum(Cal3), numpy.mean(Cal3), numpy.std(Cal3)

def OffProb(w,Station,k,StopNum): #調整下車機率
    past = sum(i <= w for i in Station)
    if (StopNum+3-w-past-1) <=0:
        basic = [1]*StopNum
    else:
        basic = [1/(StopNum+3-w-past-1)]*StopNum
    for i in range(w+1):
        basic[i] = 0
    for i in range(3):
        temp = Station[i]
        basic[temp] = basic[temp]*2  
    return(basic)

def DataGenerator(seed, Custom):
    Custom = list(Custom)
    random.seed(seed)
    numpy.random.seed(seed)
    #生成廣告
    Scene_AD = [10,25,50]
    ADChoice = Custom[0]
    M = Scene_AD[ADChoice]
    TimeCHOICE = [30,60] 
    AdTime={}
    AdName={}
    for w in range(M):
        AdTime[w] = random.choice(TimeCHOICE)
        if w < 26:
            AdName[chr(65+w)] = w
        else :
            AdName[chr(71+w)] = w
    #生成時間段
    Scene_SLOT = [(360,720),(720,1080),(1080,1440)]
    SlotChoice = Custom[1]
    StopNum = random.randint(20,90) #redefine
    N={}
    TotalN = random.randint(Scene_SLOT[SlotChoice][0],Scene_SLOT[SlotChoice][1]) #fixed sum
    N = numpy.random.multinomial(TotalN, numpy.ones(StopNum)/StopNum, size=1)[0]
    remain = numpy.random.multinomial(N[StopNum-1], numpy.ones(StopNum-1)/(StopNum-1), size=1)[0]
    for w in range(StopNum-1):
        N[w] = N[w]+remain[w]
    N[StopNum-1] = 0
    #生成客群
    K=12 
    AudienceNum = {}
    if Custom[2] == 0: #一種客群
        AudiProb = [0]*K
        temp = random.randint(0,11)
        AudiProb[temp] = 1
    elif Custom[2] == 1: #四種客群
        Scene_SexProb = [0.5,0.5]
        Scene_AgeProb = [0,1,0,1,0,0] #指定兩種年齡
        AudiProb = [round(x * y,4) for x, y in product(Scene_SexProb, Scene_AgeProb)]
    else: #隨機
        Scene_SexProb = [0.5,0.5]
        Scene_AgeProb = [1/6]*6
        AudiProb = [round(x * y,4) for x, y in product(Scene_SexProb, Scene_AgeProb)] 
        #模擬各個客群的大站
    On = {}
    AudiOn = {}
    StationOfCustomer=[]
    for k in range(K):
        random.seed(seed+k)
        StationOfCustomer.append(random.sample(list(range(0,StopNum-1)),3))
        totalOnBus=[]
    for w in range(StopNum-1):
        cusType=[]
        totalOn=random.randint(0,10)
        totalOnBus.append(totalOn)#紀錄各站上來的人數
        for p in range(totalOn):
            cusType.append(random.choices(range(K),AudiProb)) #紀錄cusType的客群人數
        for k in range(K): #紀錄w,k
            On[w,k]=cusType.count([k])
    for k in range(K):
        for w in range(StopNum-1):
            CurrOff = OffProb(w,StationOfCustomer[k],k,StopNum)
            for p in range(On[w,k]):
                offstation = random.choices(range(StopNum),CurrOff)
                AudiOn[w,k,p] = (offstation[0])
        #處理最後的AudienceNum
    for w in range(StopNum):
        for k in range(K):
            AudienceNum[w,k] = 0
    for w in range(StopNum-1):
        for k in range(K):#處理上車
            if(w==0):
                AudienceNum[w,k]=AudienceNum[w,k]+On[w,k]
            else:
                AudienceNum[w,k]=AudienceNum[w-1,k]+On[w,k]
            ##把在該站前的人都traverse一遍看誰要下車
        for g in range(w):#在站w前
            for j in range(K):#看看每個客群
                for p in range(0,On[g,j]):
                    if (AudiOn[g,j,p]==w):
                        AudienceNum[w,j]=AudienceNum[w,j]-1
    #生成喜好度
    K_Des = []
    if Custom[3] == 0: #Non-Binary, 全部1-5隨機
        DifferChoice = [1]*K
    elif Custom[3] == 1: #Non-Binary, 隨機的差異偏好
        DifferChoice = [0]*K
        for k in range(K):
            K_Des.append(list(numpy.random.randint(0,2,size=M)))
    elif Custom[3] == 2: #有些人就是討厭所有的廣告，有些人就是所有廣告都喜歡看
        DifferChoice = [0]*K
        for k in range(K):
            temp = random.randint(0,1)
            K_Des.append([temp]*M)
    else: #有些廣告就是所有客群都喜歡，有些廣告就是所有人都討厭看
        DifferChoice = [0]*K
        temp = list(numpy.random.randint(0,2, size=M))
        for k in range(K):
            K_Des.append(temp)
    AudiencePreference = {}
    for k in range(K):
        for m in range(M):
            if DifferChoice[k] == 0: #big
                if K_Des[k][m] == 0:
                    AudiencePreference[m,k] = random.randint(1,2)
                else:
                    AudiencePreference[m,k] = random.randint(4,5)
            else: #general
                AudiencePreference[m,k] = random.randint(1,5)
    #生成conflict
    Conflict = {}
    Scene_Conflict = [0, 0.1, 0.25]
    CON = Custom[4]
    for j1 in range(M):
        for j2 in range(j1,M):
            Conflict[j1,j2] = random.choices([0,1],[Scene_Conflict[CON],1-Scene_Conflict[CON]])[0] #Scene_Conflict[CON]的機率會有衝突
            Conflict[j2,j1] = Conflict[j1,j2]
    result = OrderedDict(sorted(Conflict.items(), key = lambda x: x[0][0]))
    Conflict = dict(result)
    #開跑
    TickSize=10
    TotalN = 0
    TotalBusTime = {}
    TotalBusTime[0] = 0
    for w in range(StopNum):
        TotalN = TotalN + N[w]
        TotalBusTime[w+1] = TotalN
    TotalN = TotalN + N[StopNum-1]
    TotalBusTime[StopNum] = TotalN
    for i in range(M):
        AdTime[i] = int(AdTime[i]/TickSize)
    #呼叫三種算法
    Take_Obj, Take_Sum, Take_Mean, Take_Std = TakeTurn(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum)
    Rand_Obj, Rand_Sum, Rand_Mean, Rand_Std = RandomTurn(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum, seed)
    Our_Obj, Our_Gap, Duration, Our_Sum, Our_Mean, Our_Std = OurAlgorithm(StopNum, K,N,M, TotalN, AdTime,AudienceNum, AudiencePreference,TotalBusTime,Conflict)
    ans = (M,chr(Custom[1]+65), chr(Custom[2]+65), chr(Custom[3]+65), chr(Custom[4]+65), Take_Obj,Take_Sum, Take_Mean, Take_Std, Rand_Obj,Rand_Sum, Rand_Mean, Rand_Std,Our_Obj,Our_Sum, Our_Mean, Our_Std, Our_Gap, Duration, seed)
    return ans
    
    
for i in range(1):
    for j in range(10):
        seed = random.randint(1,10000)
        data = DataGenerator(seed,custom[i])
        sheet.append(data)

wb.save(path)

        


# In[4]:


custom.index([0,0,2,4,2])

