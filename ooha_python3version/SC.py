
# coding: utf-8

# In[ ]:


import random
import math
import numpy
import dill
from itertools import product
from collections import OrderedDict
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os

fold = os.getcwd()
path = fold + "/realrundata_sc.xlsx"
column_Name = ('Min_SC','Sum_SC','Time')
if os.path.exists("realrundata_sc.xlsx"):
    wb = load_workbook(filename= path)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active
    sheet.append(column_Name)
    
re = []

def Schedule(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum, Conflict):
    
    #依目標值排序
    def sort_by_obj(ad,score,M):
        for j in range(M):
            for k in range(j+1,M):
                if(score[j]>score[k]):
                    t = score[j]
                    score[j] = score[k]
                    score[k] = t
                    g = ad[j]
                    ad[j] = ad[k]
                    ad[k] = g
        return ad
    
    #檢查是否遵守排序規則
    def valid(y,TotalN,K,M,AdTime, Conflict):
        AdRepeatTime = 10
        AdRepeatNum = 10
        AdContinousNum = 3
        ConflictNum = 5

        previous = -1 #上個廣告
        continuous = 0 #重複次數
        indi_i = 0;
        for i in range(TotalN):
            if i == indi_i:
                #重複播放限制
                if y[i]==previous:
                    continuous += 1
                else:
                    continuous = 1
                if continuous > AdContinousNum:
                    #print('重複')
                    return False
                previous = y[i]
                #一段時間重複播放限制
                counter = 0
                for j in range(1,AdTime[int(y[i])]*AdRepeatTime):
                    if i+j>=TotalN:
                        break
                    if y[i+j] == y[i]:
                        counter +=1
                if counter > AdTime[int(y[i])]*AdRepeatNum:
                    #print('時間段重複')
                    return False
                #衝突限制
                for j in range(1,ConflictNum):
                    if i+j>=TotalN:
                        break
                    if Conflict[int(y[i]),int(y[i+j])] == 0:
                        #print('衝突')
                        return False
                indi_i += AdTime[int(y[i])]
        return True

    #轉換時段
    num = []
    permue = list(range(TotalN))
    TimeSlot = {}
    counter = 0
    for i in range(TotalN):
        if i<TotalBusTime[counter]:
            for k in range(12):
                TimeSlot[i,k] = AudienceNum[counter,k]
        else:
            counter+=1
            for k in range(K):
                TimeSlot[i,k] = AudienceNum[counter,k]

    """
    #排序時段
    for i in range(TotalN):
        num.append(sum(TimeSlot[i,k] for k in range(K)))
    permue = sort_by_obj(permue,num,TotalN)
    permue.reverse()
    """    

    #初始化各廣告分數和時段是否已使用
    score = []
    occupied = []
    for i in range(M):
        score.append(int(0))
    for i in range(TotalN):
        occupied.append(int(0))
    
    #開始分時段給各廣告
    y = numpy.zeros(TotalN)
    counter = 0
    for i in range(TotalN):
        #已占用則跳過
        if occupied[permue[i]] == 1:
            continue
        #排序目前的曝光度
        ad_unsorted = list(range(M))
        tmp = []
        for j in range(M):
            tmp.append(score[j])
        ad = sort_by_obj(ad_unsorted,tmp,M)
        #決定廣告
        dec = 0
        for j in range(M):
            tmp = counter
            for k in range(AdTime[ad[j]]):
                if i+k < TotalN:
                    y[permue[i+k]] = ad[j]
                    counter = counter + 1
            if valid(y,counter,K,M,AdTime, Conflict):
                dec = j
                break
            else:
                counter = tmp
        #計算曝光度
        value = 0
        for k in range(K):
            value = value + TimeSlot[permue[i],k]*AudiencePreference[ad[dec],k]
        score[ad[dec]] = score[ad[dec]] + value
        
        #占用時段
        for j in range(AdTime[ad[dec]]):
            if i+j < TotalN:
                occupied[permue[i+j]] = 1
                y[permue[i+j]] = ad[dec]
    
    return min(score) , sum(score[i] for i in range(M))
        
        

def DataLoader(caseNum):
    file = open('data','rb')
    data = dill.load(file)
    timeResult = data[caseNum][12]
    #StopNum：站數
    StopNum = len(timeResult)
    #Nw：第w站所花時間段數量，maxN：各站時間段總數
    N = {}
    temp = 0
    TotalN = 0
    TickSize=10
    TotalBusTime = {}
    TotalBusTime[0] = 0
    for w in range(StopNum-1):
        N[w] = math.ceil((timeResult[w+1]-timeResult[w]-temp)/TickSize)
        TotalN = TotalN + N[w]
        TotalBusTime[w+1] = TotalN
        temp = TotalBusTime[w+1]*TickSize-timeResult[w+1]
    N[StopNum-1] = 0
    TotalN = TotalN + N[StopNum-1]
    TotalBusTime[StopNum] = TotalN

    at = load_workbook('adtime.xlsx')
    #AdTime j: 第j種廣告時長會佔據多少時間段
    AdTime = {}
    AdName = {}
    for j in range(at['1'].max_row-1):
        AdTime[j] = at['1']['B'][j+1].value
        AdName[at['1']['A'][j+1].value] = j
    #M：廣告數量
    M = at['1'].max_row-1
    #AdRepeatTime, AdRepeatNum: 最短P個時間段內不可播放同樣廣告超過K次
    AdRepeatTime = 10
    AdRepeatNum = 10
    #AdContinousNum: 不可連續撥放AdContinousNum次
    AdContinousNum = 5

    #Conflict j1j2, ConflictNum: 第j1和第j2個廣告可同路線ConflictNum次內播放(1:可    0:不可)
    Conflict = {}
    for j1 in range(M):
        for j2 in range(M):
            Conflict[j1,j2] = 1
    cft = load_workbook('conflict.xlsx')
    for i in range(cft['1'].max_row):
        j1 = AdName.get(cft["1"]["A"][i].value)
        j2 = AdName.get(cft["1"]["B"][i].value)
        Conflict[j1,j2] = 0
        Conflict[j2,j1] = 0

    ConflictNum = 5
    #AudiencePreference jk: 客群k喜歡廣告j的程度 
    AudiencePreference = {}
    pre = load_workbook('preference.xlsx')
    for j in range(M):
        temp = 0
        for k in pre['1'][j+2]:
            if temp == 0:
                j1 = AdName.get(k.value)
                temp = temp+1
            else:
                AudiencePreference[j1,temp-1] = k.value 
                temp = temp+1
    #K: 客群種類數
    K = temp-1
    passengerResults = {}
    for i in range(12):
        passengerResults[i] = data[caseNum][i]
    #AudienceNum w,k: w時段車上有幾位屬於客群k的人
    AudienceNum = {}
    for k in range(K):
        for w in range(StopNum):
            AudienceNum[w,k] = data[caseNum][k][w]
        
    for i in range(M):
        AdTime[i] = int(AdTime[i]/TickSize)
    tStart = time.time()
    SC_obj,SC_sum = Schedule(TotalN,K,M,AdTime,AudienceNum, AudiencePreference, TotalBusTime,StopNum,Conflict)
    tEnd = time.time()
    ans = (SC_obj,SC_sum,tEnd-tStart)
    print('最小曝光度:',SC_obj,'總合曝光度:',SC_sum)
    return ans

for i in range(0,100): #改這邊的range，上限99
    data = DataLoader(i)
    sheet.append(data)
    wb.save(path)
